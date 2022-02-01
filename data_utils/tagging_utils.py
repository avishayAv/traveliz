import pickle as pkl
from datetime import datetime
from typing import Optional

import numpy
import pandas
import pandas as pd
from openpyxl import load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.formatting.rule import Rule
from openpyxl.styles import PatternFill, Font
from openpyxl.styles.differential import DifferentialStyle
from styleframe import StyleFrame

from ParsingFunctions import searching_for_sublet
from PriceParser import Price
from Sublet import Rooms
from unit_tests.TestsDB import TestGroundTruth, TestRawInput, Test
from utils import whatsapp_group_to_location


def create_pkl_for_test():
    a = pkl.load(open("dict_of_sublets.p", "rb"))
    data_for_tests = []
    for group_id, posts in a.items():
        for post in posts:
            if len(post['post_text']) > 20 and not searching_for_sublet(post.get('post_title', ''), post['post_text']):
                post['group_id'] = group_id
                data_for_tests.append(post)
    random_400_posts_from_facebook = numpy.random.choice(data_for_tests, size=min(400, len(data_for_tests)),
                                                         replace=False, p=None)
    pkl.dump(random_400_posts_from_facebook, open("data_utils/random_facebook_posts.pkl", "wb"))


def if_sheet_name_exists(excel_path:str,
                         sheet_name : str):
    xl = pd.ExcelFile(excel_path)
    return sheet_name in xl.sheet_names


def create_excel_for_facebook_data(pkl_path: str = "random_facebook_posts.pkl",
                                   excel_path: str = "data_for_tagging.xlsx",
                                   sheet_name: str = "facebook"):
    assert not if_sheet_name_exists(excel_path, sheet_name), f"sheet name {sheet_name} already exists in {excel_path}"
    posts = pkl.load(open(pkl_path, "rb"))

    text = [x['post_text'] for x in posts]
    post_id = [x['post_id'] for x in posts]
    start_date = ["start_date" for _ in post_id]
    end_date = ["end_date" for _ in post_id]
    price = ["price" for _ in post_id]
    location = ["location" for _ in post_id]
    phone_number = ["phone_number" for _ in post_id]
    rooms = ["rooms" for _ in post_id]
    # Create some Pandas dataframes from some data.
    df1 = pd.DataFrame({'Post_id': post_id,
                        'Text': text,
                        'Start_date': start_date,
                        'End_date': end_date,
                        'Price': price,
                        'Location': location,
                        'Phone_number': phone_number,
                        "Rooms": rooms})

    book = load_workbook(excel_path)
    writer = pd.ExcelWriter(excel_path, engine='openpyxl')
    writer.book = book
    StyleFrame(df1).to_excel(writer, sheet_name=sheet_name).save()
    writer.close()
    file_path = excel_path

    index_rules = str(len(posts) + 1)
    add_rules_to_excel(file_path, index_rules, sheet_name)


def create_excel_for_whatsapp_data(pkl_path: str = "sublets_from_whatsapp.p",
                                   excel_path: str = "data_for_tagging.xlsx",
                                   sheet_name: str = "whatsapp"):
    assert not if_sheet_name_exists(excel_path, sheet_name), f"sheet name {sheet_name} already exists in {excel_path}"
    posts = pkl.load(open(pkl_path, "rb"))
    text, dates, sender_phone, group_name, start_date, end_date, price, location, location, phone_number, rooms = \
        [], [], [], [], [], [], [], [], [], [], [],
    for key in posts.keys():
        posts_by_group = posts[key]
        text += [x[0] for x in posts_by_group]
        dates += [x[1] for x in posts_by_group]
        sender_phone += [x[2].phone for x in posts_by_group]
        group_name += [key for _ in posts_by_group]
        start_date += ["start_date" for _ in posts_by_group]
        end_date += ["end_date" for _ in posts_by_group]
        price += ["price" for _ in posts_by_group]
        location += ["location" for _ in posts_by_group]
        phone_number += ["phone_number" for _ in posts_by_group]
        rooms += ["rooms" for _ in posts_by_group]

    # Create some Pandas dataframes from some data.
    df1 = pd.DataFrame({'Group_name/Post_date/Post_phone': [x + '\n' + str(y) + '\n' + str(z).replace('+972', '') for
                                                            x, y, z in zip(group_name, dates, sender_phone)],
                        'Text': text,
                        'Start_date': start_date,
                        'End_date': end_date,
                        'Price': price,
                        'Location': location,
                        'Phone_number': phone_number,
                        "Rooms": rooms})

    book = load_workbook(excel_path)
    writer = pd.ExcelWriter(excel_path, engine='openpyxl')
    writer.book = book
    StyleFrame(df1).to_excel(writer, sheet_name=sheet_name).save()
    writer.close()
    index_rules = str(len(text) + 1)
    add_rules_to_excel(excel_path, index_rules, sheet_name)


def add_rules_to_excel(file_path:str,
                       index_rules: str,
                       sheet_name: str
                       ):
    red_text = Font(color="9C0006")
    red_fill = PatternFill(bgColor="FFC7CE")
    dxf = DifferentialStyle(font=red_text, fill=red_fill)
    # if cell contains orig text fill in with red color
    rule1 = Rule(type="containsText", operator="containsText", text="start_date", dxf=dxf)
    rule2 = Rule(type="containsText", operator="containsText", text="end_date", dxf=dxf)
    rule3 = Rule(type="containsText", operator="containsText", text="price", dxf=dxf)
    rule4 = Rule(type="containsText", operator="containsText", text="location", dxf=dxf)
    rule5 = Rule(type="containsText", operator="containsText", text="phone_number", dxf=dxf)
    rule6 = Rule(type="containsText", operator="containsText", text="rooms", dxf=dxf)

    rule1.formula = ['NOT(ISERROR(SEARCH("start_date",C2)))']
    rule2.formula = ['NOT(ISERROR(SEARCH("end_date",D2)))']
    rule3.formula = ['NOT(ISERROR(SEARCH("price",E2)))']
    rule4.formula = ['NOT(ISERROR(SEARCH("location",F2)))']
    rule5.formula = ['NOT(ISERROR(SEARCH("phone_number",G2)))']
    rule6.formula = ['NOT(ISERROR(SEARCH("rooms",H2)))']

    # if cell is empty fill in with yellow color
    yellowFill = PatternFill(start_color='00FFFF00', end_color='00FFFF00', fill_type='solid')
    rule_7 = FormulaRule(formula=['ISBLANK(C1)'], stopIfTrue=True, fill=yellowFill)
    rule_8 = FormulaRule(formula=['ISBLANK(D1)'], stopIfTrue=True, fill=yellowFill)
    rule_9 = FormulaRule(formula=['ISBLANK(E1)'], stopIfTrue=True, fill=yellowFill)
    rule_10 = FormulaRule(formula=['ISBLANK(F1)'], stopIfTrue=True, fill=yellowFill)
    rule_11 = FormulaRule(formula=['ISBLANK(G1)'], stopIfTrue=True, fill=yellowFill)
    rule_12 = FormulaRule(formula=['ISBLANK(H1)'], stopIfTrue=True, fill=yellowFill)

    wb = load_workbook(file_path)
    ws = wb[sheet_name]
    ws.conditional_formatting.add(f'C2:C{index_rules}', rule1)
    ws.conditional_formatting.add(f'D2:D{index_rules}', rule2)
    ws.conditional_formatting.add(f'E2:E{index_rules}', rule3)
    ws.conditional_formatting.add(f'F2:F{index_rules}', rule4)
    ws.conditional_formatting.add(f'G2:G{index_rules}', rule5)
    ws.conditional_formatting.add(f'H2:H{index_rules}', rule6)
    ws.conditional_formatting.add(f'C1:C{index_rules}', rule_7)
    ws.conditional_formatting.add(f'D1:D{index_rules}', rule_8)
    ws.conditional_formatting.add(f'E1:E{index_rules}', rule_9)
    ws.conditional_formatting.add(f'F1:F{index_rules}', rule_10)
    ws.conditional_formatting.add(f'G1:G{index_rules}', rule_11)
    ws.conditional_formatting.add(f'H1:H{index_rules}', rule_12)

    # as default set cells colors to green
    for column in [ws['C'], ws['D'], ws['E'], ws['F'], ws['G'], ws['H']]:
        for cell in column:
            cell.fill = PatternFill(start_color='0000FF00', end_color='0000FF00', fill_type='solid')

    ws['C1'].fill = PatternFill(start_color='00FFFFFF', end_color='00FFFFFF', fill_type='solid')
    ws['D1'].fill = PatternFill(start_color='00FFFFFF', end_color='00FFFFFF', fill_type='solid')
    ws['E1'].fill = PatternFill(start_color='00FFFFFF', end_color='00FFFFFF', fill_type='solid')
    ws['F1'].fill = PatternFill(start_color='00FFFFFF', end_color='00FFFFFF', fill_type='solid')
    ws['G1'].fill = PatternFill(start_color='00FFFFFF', end_color='00FFFFFF', fill_type='solid')
    ws['H1'].fill = PatternFill(start_color='00FFFFFF', end_color='00FFFFFF', fill_type='solid')
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 100
    ws.column_dimensions['B'].heigth = 400
    for column in [ws.column_dimensions['C'], ws.column_dimensions['D'], ws.column_dimensions['E'],
                   ws.column_dimensions['F'], ws.column_dimensions['G'], ws.column_dimensions['H']]:
        column.width = 20

    wb.save(file_path)


def read_excel_and_create_tagged_df(file_name: str = "data_for_tagging.xlsx",
                                    name: Optional[str] = None):
    df = read_excel(file_name, name=name)
    for sheet_name, sheet_df in df.items():
        yield sheet_name, sheet_df


def read_excel(file_name: str, name: str):
    return pandas.read_excel(file_name, sheet_name=name)


# TODO [ES] : create some more tabs for whatsapp samples
def create_tests_from_tagged_excel():
    def add_zero(x):
        if x[0] != '0':
            return '0' + x
        return x

    def date_str_to_datetime(date_str):
        if type(date_str) == datetime:
            date_str = date_str.strftime('%m/%d/%Y')
        if date_str is None:
            return None
        date_str = date_str.replace('.', '/')
        if '/' not in date_str:
            return None
        d, m, y = date_str.split('/')
        d = '0' + d if len(d) == 1 else d
        m = '0' + m if len(m) == 1 else m
        y = '20' + y if len(y) == 2 else y
        assert len(d) == 2 and len(m) == 2 and len(y) == 4
        return datetime.strptime('/'.join([d, m, y]), '%d/%m/%Y').date()

    tests_db = []
    posts = pkl.load(open("../data_utils/random_facebook_posts.pkl", "rb"))
    post_id_to_post = {str(x['post_id']): x for x in posts}
    for data_source, tagged_data_df in read_excel_and_create_tagged_df():
        tagged_data_df = tagged_data_df.where(pd.notnull(tagged_data_df), None)

        for _, tagged_item in tagged_data_df.iterrows():
            price = Price()
            prices = tagged_item['Price']
            if prices is None or 'price' in prices:
                prices = None
            else:
                for p in prices.split(','):
                    amount, description = p.split()
                    if description == 'd':
                        price.price_per_night = int(amount)
                    elif description == 'e': # weekend
                        price.price_per_weekend = int(amount)
                    elif description == 'w':
                        price.discounted_price_per_night = int(amount) // 7
                        price.discounted_period = 7
                    elif description == 'm':
                        price.price_per_month = int(amount)
                    else: # period
                        price.discounted_price_per_night = int(amount) // int(description)
                        price.discounted_period = int(description)
            phone_number = str(tagged_item['Phone_number']).replace(' ', '').replace('-', '')
            if phone_number is None or 'phone' in phone_number:
                phone_number = None
            elif phone_number.isnumeric():
                phone_number = {add_zero(phone_number)}
            else:
                phone_number = set([add_zero(x) for x in phone_number.split(',')])

            start_date = date_str_to_datetime(tagged_item['Start_date'])
            end_date = date_str_to_datetime(tagged_item['End_date'])
            rooms = Rooms()
            raw_rooms = tagged_item['Rooms']
            if type(raw_rooms) in [int, float]:
                rooms.number = float(raw_rooms)
            else:
                split_raw_rooms = raw_rooms.split()
                if len(split_raw_rooms) > 1:
                    rooms.number = float(split_raw_rooms[0]) if split_raw_rooms[0] != 'None' else None
                    rooms.shared = True if split_raw_rooms[1] == 's' else False
                else:
                    rooms.number = None
            location = tagged_item['Location']
            if 'location' in location:
                location = None
            gt = TestGroundTruth(start_date=start_date,
                                 end_date=end_date,
                                 price=price, location=location,
                                 phone_number=phone_number, rooms=rooms)
            if data_source == 'facebook':
                post = post_id_to_post[str(tagged_item['Post_id'])]
                raw_input = TestRawInput(group_id=post['group_id'], location=post.get('location'),
                                         post_id=post['post_id'], price=post.get('price'),
                                         text=post['text'], title=post.get('title', ''),
                                         post_time=post['time'])
            else:
                assert data_source == 'whatsapp'
                # self, post_title='', post_text='', post_time=None, listing_location=None, group_id=None
                # post_text=message['text'], post_time=post_time, listing_location=group_to_location[group_name]
                # listing_price=None
                group_name, post_date, sender_phone = tagged_item['Group_name/Post_date/Post_phone'].split('\n')
                raw_input = TestRawInput(location=whatsapp_group_to_location[group_name],
                                         text=tagged_item['Text'],
                                         post_time=datetime.strptime(post_date, '%Y-%m-%d'),
                                         phone_number=sender_phone)
            test = Test(gt=gt, raw_input=raw_input, source=data_source)
            if test.is_test_tagged():
                tests_db.append(test)
    pkl.dump(tests_db, open('../unit_tests/test_db.p', 'wb'))


if __name__ == '__main__':
    # create_pkl_for_test()
    # create_excel_for_whatsapp_data()
    create_tests_from_tagged_excel()
